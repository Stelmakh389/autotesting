from django.shortcuts import get_object_or_404, redirect, render
from django.db import models  # Добавляем этот импорт
from django.db.models import Q
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.urls import reverse_lazy
from django.contrib.messages.views import SuccessMessageMixin
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from .models import Equipment, EquipmentGroup
from .forms import EquipmentForm, EquipmentGroupForm
from django.contrib import messages
from django.urls import reverse
from datetime import datetime
import openpyxl
from datetime import datetime
import io
from django.http import HttpResponse
from .forms import CSVImportForm
import json
from django.http import JsonResponse
from django.db.models import ExpressionWrapper, F, DateField
from django.db.models.functions import Now


# Базовые миксины для оборудования
class BaseEquipmentMixin:
    template_name_suffix = None
    equipment_type_name = None
    equipment_type_name_accusative = None
    url_prefix = None
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'equipment_type_name': self.equipment_type_name,
            'equipment_type_name_accusative': self.equipment_type_name_accusative,
            'create_url': f'equipment:equipment-create',  # Добавляем namespace
            'update_url': f'equipment:equipment-update',  # Добавляем namespace
            'delete_url': f'equipment:equipment-delete',  # Добавляем namespace
            'cancel_url': reverse_lazy(f'equipment:{self.url_prefix}-list'),  # Добавляем namespace
        })
        return context

class BaseEquipmentListView(LoginRequiredMixin, BaseEquipmentMixin, ListView):
    template_name = 'equipment/equipment_list.html'
    context_object_name = 'equipment_list'
    paginate_by = 30
    ordering = ['-id']

    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Фильтрация по типу
        equipment_type = self.request.GET.get('type')
        if equipment_type:
            queryset = queryset.filter(equipment_type=equipment_type)
            
        # Сортировка
        sort = self.request.GET.get('sort')
        order = self.request.GET.get('order')
        if sort:
            if order == 'desc':
                sort = f'-{sort}'
            queryset = queryset.order_by(sort)
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'search_query': self.request.GET.get('search', ''),
        })
        return context

class BaseEquipmentCreateView(LoginRequiredMixin, SuccessMessageMixin, BaseEquipmentMixin, CreateView):
    template_name = 'equipment/equipment_form.html'
    
    def get_success_url(self):
        return reverse_lazy(f'equipment:{self.url_prefix}-list')

class BaseEquipmentUpdateView(LoginRequiredMixin, SuccessMessageMixin, BaseEquipmentMixin, UpdateView):
    template_name = 'equipment/equipment_form.html'
    
    def get_success_url(self):
        return reverse_lazy (f'equipment:{self.url_prefix}-list')

class BaseEquipmentDeleteView(LoginRequiredMixin, SuccessMessageMixin, BaseEquipmentMixin, DeleteView):
    template_name = 'equipment/equipment_confirm_delete.html'
    
    def get_success_url(self):
        return reverse_lazy(f'equipment:{self.url_prefix}-list')
    
class EquipmentDetailView(LoginRequiredMixin, DetailView):
   model = Equipment
   template_name = 'equipment/equipment_detail.html'
   context_object_name = 'equipment'
   
   def get_context_data(self, **kwargs):
       context = super().get_context_data(**kwargs)
       context['fields'] = self.model._meta.fields
       return context

class EquipmentListView(BaseEquipmentListView):
    model = Equipment
    equipment_type_name = 'Оборудование' 
    equipment_type_name_accusative = 'оборудование'
    url_prefix = 'equipment'
    template_name = 'equipment/equipment_list.html'

    default_columns = [
        'equipment_type', 
        'name', 
        'zav_nomer', 
        'inv_nomer', 
        'reg_nomer', 
        'klass_toch', 
        'predel'
    ]
   
    AVAILABLE_COLUMNS = {
        'equipment_type': 'Тип',
        'name': 'Наименование',
        'tip': 'Тип',
        'zav_nomer': 'Заводской №',
        'inv_nomer': 'Инв. №', 
        'reg_nomer': 'Рег. №',
        'kol_vo': 'Кол-во',
        'klass_toch': 'Класс точности',
        'predel': 'Предел измерений',
        'period_poverk': 'Периодичность поверки',
        'category_si': 'Категория СИ',
        'organ_poverk': 'Орган поверки',
        'data_poverk': 'Дата поверки',
        'srok_poverk': 'Срок поверки',
        'other': 'Примечание'
    }

    def get_selected_columns(self):
        return self.request.session.get('equipment_columns', self.default_columns)

    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Поиск по всем полям
        search_query = self.request.GET.get('search')
        if search_query:
            # Создаем пустой Q-объект
            q_objects = Q()
            
            # Получаем все поля модели
            fields = [field.name for field in Equipment._meta.fields]
            
            # Добавляем условие поиска для каждого поля
            for field in fields:
                # Добавляем поиск по каждому полю через icontains
                q_objects |= Q(**{f"{field}__icontains": search_query})
                
            # Применяем фильтр
            queryset = queryset.filter(q_objects)
        
        # Фильтрация по типу
        equipment_type = self.request.GET.get('type')
        if equipment_type:
            queryset = queryset.filter(equipment_type=equipment_type)
            
        # Сортировка
        sort = self.request.GET.get('sort')
        if sort and sort in self.AVAILABLE_COLUMNS:
            order = self.request.GET.get('order')
            if order == 'desc':
                sort = f'-{sort}'
            queryset = queryset.order_by(sort)
        
        return queryset.distinct()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        selected_columns = self.get_selected_columns()
        
        filters = {}
        queryset = self.model.objects.all()
        
        for field in selected_columns:
            if field in self.AVAILABLE_COLUMNS:
                values = queryset.values_list(field, flat=True).distinct().order_by(field)
                filters[field] = [v for v in values if v]
                
        context.update({
            'available_columns': self.AVAILABLE_COLUMNS,
            'selected_columns': selected_columns,
            'filters': filters,
            'current_filters': {
                field: self.request.GET.get(field) 
                for field in selected_columns
            }
        })
        return context
   
@login_required 
def save_equipment_columns(request):
   if request.method == 'POST':
       selected = request.POST.getlist('columns')
       request.session['equipment_columns'] = selected
       messages.success(request, 'Настройки отображения сохранены')
   return redirect('equipment:equipment-list')

class EquipmentCreateView(BaseEquipmentCreateView):
    model = Equipment
    form_class = EquipmentForm
    equipment_type_name = 'Оборудование'
    equipment_type_name_accusative = 'оборудование'
    url_prefix = 'equipment'
    success_message = "Элемент успешно создан"

class EquipmentUpdateView(BaseEquipmentUpdateView):
    model = Equipment
    form_class = EquipmentForm
    equipment_type_name = 'Оборудование'
    equipment_type_name_accusative = 'оборудование' 
    url_prefix = 'equipment'
    success_message = "Элемент успешно обновлен"

class EquipmentDeleteView(BaseEquipmentDeleteView):
    model = Equipment
    equipment_type_name = 'Оборудование'
    equipment_type_name_accusative = 'оборудование'
    url_prefix = 'equipment'
    success_message = "Элемент успешно удален"

@login_required
def duplicate_equipment(request, pk, equipment_type):
    """Дублирование записи оборудования с учётом типа"""
    source_equipment = get_object_or_404(Equipment, pk=pk)
    source_equipment.pk = None  # Обнуляем первичный ключ для создания нового объекта
    source_equipment.name = f"{source_equipment.name} (копия)"  # Добавляем пометку
    source_equipment.save()
    
    messages.success(request, f"Элемент успешно скопирован")
    return redirect(f'equipment:{equipment_type}-list')

@login_required
def delete_all_equipment(request, equipment_type):
    equipment_type_mapping = {
        'measurement': ('measurement_tool', 'средств измерений', 'measurement-tool-list'),
        'testing': ('testing_equipment', 'испытательного оборудования', 'testing-equipment-list'),
        'auxiliary': ('auxiliary_equipment', 'вспомогательного оборудования', 'auxiliary-equipment-list'),
    }

    if equipment_type not in equipment_type_mapping:
        messages.error(request, 'Неверный тип оборудования')
        return redirect('equipment:equipment-list')

    equipment_type_value, equipment_type_name, redirect_url_name = equipment_type_mapping[equipment_type]

    if request.method == 'POST':
        # Удаляем все записи данного типа
        deleted_count = Equipment.objects.filter(equipment_type=equipment_type_value).delete()[0]
        messages.success(request, f'Удалено записей: {deleted_count}')
        return redirect(redirect_url_name)

    # Получаем количество записей для отображения
    items_count = Equipment.objects.filter(equipment_type=equipment_type_value).count()

    return render(request, 'equipment/delete_all_confirm.html', {
        'equipment_type_name': equipment_type_name,
        'items_count': items_count,
        'cancel_url': redirect_url_name
    })


@login_required
def import_equipment(request):
   if request.method == 'POST':
       form = CSVImportForm(request.POST, request.FILES)
       if form.is_valid():
           try:
               excel_file = request.FILES['csv_file']
               wb = openpyxl.load_workbook(excel_file)
               ws = wb.active
               
               fields = [f.name for f in Equipment._meta.fields]
               stats = {'updated': 0, 'created': 0, 'skipped': 0}
               
               headers = {cell.value: idx for idx, cell in enumerate(ws[1])}
               
               for row in ws.iter_rows(min_row=2):
                   try:
                       cleaned_data = {}
                       
                       # Получаем ID из строки Excel
                       item_id = row[headers.get('id', 0)].value
                       
                       for field in fields:
                           if field in headers and field != 'id':
                               value = row[headers[field]].value
                               if value:
                                   if isinstance(value, str):
                                       value = value.strip()
                                   if field in ['data_poverk', 'srok_poverk']:
                                       if isinstance(value, datetime):
                                           value = value.date()
                               cleaned_data[field] = value

                       if item_id:
                           # Если есть ID, обновляем существующую запись
                           equipment = Equipment.objects.filter(id=item_id).first()
                           if equipment:
                               for field, value in cleaned_data.items():
                                   setattr(equipment, field, value)
                               equipment.save()
                               stats['updated'] += 1
                           else:
                               Equipment.objects.create(**cleaned_data)
                               stats['created'] += 1
                       else:
                           Equipment.objects.create(**cleaned_data)
                           stats['created'] += 1

                   except Exception as e:
                       print(f"Row processing error: {e}")
                       stats['skipped'] += 1

               messages.success(request, 
                   f'Импорт завершен: обновлено {stats["updated"]}, '
                   f'создано {stats["created"]}, пропущено {stats["skipped"]} записей')

           except Exception as e:
               messages.error(request, f'Ошибка при импорте: {str(e)}')
           
           return redirect('equipment:equipment-list')
   else:
       form = CSVImportForm()

   example_header = ','.join([f.name for f in Equipment._meta.fields])
   
   return render(request, 'equipment/import_csv.html', {
       'form': form,
       'example_header': example_header
   })

@login_required 
def export_equipment(request):
   wb = openpyxl.Workbook()
   ws = wb.active
   
   fields = [field.name for field in Equipment._meta.fields]
   ws.append(fields)

   items = Equipment.objects.values_list(*fields)
   for item in items:
       ws.append(['' if value is None else value for value in item])

   response = HttpResponse(
       content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
   )
   response['Content-Disposition'] = 'attachment; filename="equipment.xlsx"'
   wb.save(response)
   
   return response

@login_required
def bulk_delete_equipment(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        ids = data.get('ids', [])
        
        # Удаляем выбранные записи
        deleted_count = Equipment.objects.filter(id__in=ids).delete()[0]
        
        messages.success(request, f'Удалено элементов: {deleted_count}')
        return JsonResponse({'status': 'success'})
    
    return JsonResponse({'status': 'error'}, status=400)

@login_required
def bulk_duplicate_equipment(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        ids = data.get('ids', [])
        
        duplicated_count = 0
        for equipment_id in ids:
            try:
                equipment = Equipment.objects.get(id=equipment_id)
                equipment.pk = None
                equipment.name = f"{equipment.name} (копия)"
                equipment.save()
                duplicated_count += 1
            except Equipment.DoesNotExist:
                continue
        
        messages.success(request, f'Скопировано элементов: {duplicated_count}')
        return JsonResponse({'status': 'success'})
    
    return JsonResponse({'status': 'error'}, status=400)

class BaseGroupMixin:
    model = EquipmentGroup
    success_url = reverse_lazy('equipment-group-list')

class EquipmentGroupListView(LoginRequiredMixin, ListView):
    model = EquipmentGroup
    template_name = 'equipment/equipment_group_list.html'
    context_object_name = 'groups'  # Изменено с equipment_list на groups
    paginate_by = 30

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['available_columns'] = {
            'name': 'Название',
            'description': 'Описание',
            'equipment_count': 'Количество оборудования'
        }
        return context

class EquipmentGroupCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = EquipmentGroup
    form_class = EquipmentGroupForm
    template_name = 'equipment/equipment_group_form.html'
    success_url = reverse_lazy('equipment-group-list')
    success_message = "Группа оборудования успешно создана"

class EquipmentGroupUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = EquipmentGroup
    form_class = EquipmentGroupForm
    template_name = 'equipment/equipment_group_form.html'
    success_url = reverse_lazy('equipment-group-list')
    success_message = "Группа оборудования успешно обновлена"

class EquipmentGroupDeleteView(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = EquipmentGroup
    template_name = 'equipment/equipment_group_confirm_delete.html'
    success_url = reverse_lazy('equipment-group-list')
    success_message = "Группа оборудования успешно удалена"

class EquipmentGroupDetailView(LoginRequiredMixin, DetailView):
    model = EquipmentGroup
    template_name = 'equipment/equipment_group_detail.html'
    context_object_name = 'group'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Добавляем дополнительный контекст, если необходимо
        return context
    
class DeleteMixin:
    def delete(self, request, *args, **kwargs):
        object = self.get_object()
        object_name = str(object)
        object.delete()
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'status': 'success',
                'message': f'Элемент "{object_name}" успешно удален'
            })
        
        messages.success(request, f'Элемент "{object_name}" успешно удален')
        return redirect(self.get_success_url())


class EquipmentGroupDeleteView(DeleteMixin, LoginRequiredMixin, DeleteView):
    pass