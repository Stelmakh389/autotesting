from django.shortcuts import get_object_or_404, redirect, render
from django.db.models import F, DateField, IntegerField, Q
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.urls import reverse_lazy
from django.db import models
from django.contrib.messages.views import SuccessMessageMixin
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from datetime import datetime, date
import openpyxl
import json
from .models import Equipment, EquipmentGroup
from .forms import EquipmentForm, EquipmentGroupForm, CSVImportForm
from django.db.models.functions import Now
from django.core.exceptions import ObjectDoesNotExist
from django.utils.timezone import localtime
from django.template.loader import render_to_string
from django.db.models import F, IntegerField
from django.db.models.functions import ExtractDay, Cast, Now
from django.views import View
from django.core.exceptions import FieldDoesNotExist


class DeleteMixin:
    def delete(self, request, *args, **kwargs):
        object = self.get_object()
        object_name = str(object)
        object.delete()
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'status': 'success',
                'message': f'Элемент "{object_name}" успешно удален'
            })
        
        messages.success(request, f'Элемент "{object_name}" успешно удален')
        return redirect(self.get_success_url())

# Базовые миксины
class BaseEquipmentMixin:
    model = Equipment
    template_name_suffix = None
    equipment_type_name = 'Оборудование'
    equipment_type_name_accusative = 'оборудование'
    url_prefix = 'equipment'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'equipment_type_name': self.equipment_type_name,
            'equipment_type_name_accusative': self.equipment_type_name_accusative,
            'create_url': f'equipment:create',
            'update_url': f'equipment:update',
            'delete_url': f'equipment:delete',
            'cancel_url': reverse_lazy(f'equipment:list'),
        })
        return context

class BaseEquipmentListView(LoginRequiredMixin, BaseEquipmentMixin, ListView):
    template_name = 'equipment/equipment_list.html'
    context_object_name = 'equipment_list'
    paginate_by = 30
    ordering = ['-id']

    def get_queryset(self):
        # Получаем базовый queryset
        queryset = super().get_queryset()
        
        # Фильтруем по типу оборудования, если он определен
        if hasattr(self, 'equipment_type_filter'):
            queryset = queryset.filter(equipment_type=self.equipment_type_filter)
        
        # Применяем поиск, если есть
        search_query = self.request.GET.get('search')
        if search_query:
            search_fields = [field.name for field in self.model._meta.fields 
                           if isinstance(field, (models.CharField, models.TextField))]
            
            q_objects = Q()
            for field in search_fields:
                q_objects |= Q(**{f"{field}__icontains": search_query})
            
            queryset = queryset.filter(q_objects)
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['app_name'] = 'equipment'  # Указываем название приложения
        context['available_columns'] = {
            field.name: getattr(field, 'verbose_name', field.name)
            for field in Equipment._meta.get_fields()
            if isinstance(field, (models.Field, models.ForeignKey, models.ManyToManyField))
        }
        context['selected_columns'] = self.request.session.get('equipment_columns', list(context['available_columns'].keys()))
        return context
    
class BaseEquipmentCreateView(LoginRequiredMixin, SuccessMessageMixin, BaseEquipmentMixin, CreateView):
    def get_success_url(self):
        return reverse_lazy(f'equipment:list')

class BaseEquipmentUpdateView(LoginRequiredMixin, SuccessMessageMixin, BaseEquipmentMixin, UpdateView):
    model = Equipment
    template_name = 'equipment/equipment_form.html'
    context_object_name = 'equipment'
    
    def get_success_url(self):
        return reverse_lazy(f'equipment:list')

class BaseEquipmentDeleteView(LoginRequiredMixin, SuccessMessageMixin, BaseEquipmentMixin, DeleteView):
    template_name = 'equipment/equipment_confirm_delete.html'
    
    def get_success_url(self):
        return reverse_lazy(f'equipment:list')

class EquipmentDetailView(LoginRequiredMixin, DetailView):
    model = Equipment
    template_name = 'equipment/equipment_detail.html'
    context_object_name = 'equipment'

    def get(self, request, *args, **kwargs):
        try:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                instance = self.get_object()
                
                # Получаем информацию о полях модели
                fields_data = []
                for field in self.model._meta.fields:
                    fields_data.append({
                        'name': field.name,
                        'verbose_name': field.verbose_name
                    })
                
                # Получаем данные объекта
                data = {
                    'equipment_type': instance.equipment_type,
                    'equipment_type_display': instance.get_equipment_type_display(),
                    'days_between_poverk': instance.days_between_poverk,  # Добавляем дни до поверки
                    'poverk_status': instance.poverk_status,  # Добавляем статус поверки
                }
                
                # Добавляем все остальные поля
                for field in self.model._meta.fields:
                    value = getattr(instance, field.name)
                    if isinstance(value, (date, datetime)):
                        value = value.strftime('%d.%m.%Y')
                    elif hasattr(value, '__str__'):
                        value = str(value)
                    data[field.name] = value if value is not None else ''

                return JsonResponse({
                    'status': 'success',
                    'data': data,
                    'fields': fields_data
                })
            return super().get(request, *args, **kwargs)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

class EquipmentListView(LoginRequiredMixin, ListView):
    model = Equipment
    template_name = 'equipment/equipment_list.html'
    context_object_name = 'equipment_list'
    paginate_by = 30
    ordering = ['-id']

    # Автоматическое формирование списка всех полей модели
    model_fields = [field.name for field in Equipment._meta.get_fields() if isinstance(field, models.Field)]

    # Поля по умолчанию
    default_columns = ['equipment_type', 'name', 'zav_nomer', 'inv_nomer', 'reg_nomer', 'klass_toch', 'predel', 'days_until_verification']

    # Автоматическое формирование AVAILABLE_COLUMNS
    AVAILABLE_COLUMNS = {
        field.name: getattr(field, 'verbose_name', field.name)
        for field in Equipment._meta.get_fields()
        if isinstance(field, models.Field)
    }

    AVAILABLE_COLUMNS['days_until_verification'] = 'Дней до поверки'

    def get_selected_columns(self):
        """
        Получает список выбранных колонок из сессии, если они есть, иначе использует значения по умолчанию.
        """
        return self.request.session.get('equipment_columns', self.default_columns)

    def get_queryset(self):
        """
        Получает и фильтрует список оборудования с учетом поискового запроса, сортировки и фильтрации по параметрам.
        """
        queryset = super().get_queryset()

        # Аннотация количества дней до поверки (0 если уже просрочено)
        queryset = queryset.annotate(
            days_until_verification=ExtractDay(F('srok_poverk') - Now(), output_field=IntegerField())
        )

        # Фильтрация по типу оборудования
        equipment_type = self.request.GET.get('type')
        if equipment_type:
            queryset = queryset.filter(equipment_type=equipment_type)

        # Поиск по ключевым полям (можно расширить список)
        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(zav_nomer__icontains=search_query) |
                Q(inv_nomer__icontains=search_query) |
                Q(reg_nomer__icontains=search_query) |
                Q(klass_toch__icontains=search_query) |
                Q(predel__icontains=search_query)
            )

        # Сортировка
        sort = self.request.GET.get('sort')
        order = self.request.GET.get('order')
        if sort:
            if sort == 'days_until_verification':
                sort = 'days_until_verification'
            if order == 'desc':
                sort = f'-{sort}'
            queryset = queryset.order_by(sort)

        # Применяем фильтры из GET параметров
        selected_columns = self.get_selected_columns()
        for field in selected_columns:
            if field != 'days_between_poverk':  # Исключаем виртуальное поле
                value = self.request.GET.get(field)
                if value and field in self.model_fields:
                    queryset = queryset.filter(**{field: value})

        return queryset

    def get_context_data(self, **kwargs):
        """
        Добавляет в контекст данные о доступных столбцах, выбранных столбцах и фильтрах.
        """
        context = super().get_context_data(**kwargs)
        selected_columns = self.get_selected_columns()

        # Базовый queryset для фильтров
        queryset = self.model.objects.all()

        # Получаем все поля модели
        model_fields = [
            {
                'name': field.name,
                'verbose_name': field.verbose_name if hasattr(field, 'verbose_name') else field.name,
                'type': field.get_internal_type()
            }
            for field in self.model._meta.fields
        ]

        # Формируем доступные фильтры
        filters = {}
        for field in selected_columns:
            if field in self.model_fields and field != 'days_between_poverk':
                values = (
                    queryset
                    .exclude(**{f"{field}__isnull": True})
                    .values_list(field, flat=True)
                    .distinct()
                    .order_by(field)
                )
                filters[field] = [v for v in values if v]

        # Обновляем контекст
        context.update({
            'app_name': 'equipment',
            'model_fields': model_fields,
            'available_columns': self.AVAILABLE_COLUMNS,
            'selected_columns': selected_columns,
            'filters': filters,
            'Equipment': Equipment,
            'form': EquipmentForm(),
            'current_filters': {
                field: self.request.GET.get(field) for field in selected_columns
            }
        })
        return context
    
class AjaxFormMixin:
    """Миксин для обработки AJAX-запросов форм"""
    def form_valid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            self.object = form.save()
            return JsonResponse({
                'status': 'success',
                'message': 'Элемент успешно сохранен'
            })
        return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'status': 'error',
                'errors': form.errors
            })
        return super().form_invalid(form)

class EquipmentCreateView(AjaxFormMixin, BaseEquipmentCreateView):
    form_class = EquipmentForm



class EquipmentDeleteView(DeleteMixin, BaseEquipmentDeleteView):
    success_message = "Элемент успешно удален"

class EquipmentUpdateView(AjaxFormMixin, BaseEquipmentUpdateView):
    form_class = EquipmentForm

    def get(self, request, *args, **kwargs):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            instance = self.get_object()
            data = {field.name: getattr(instance, field.name) 
                   for field in self.model._meta.fields}
            # Преобразование дат в строки
            for field_name, value in data.items():
                if isinstance(value, (date, datetime)):
                    data[field_name] = value.strftime('%Y-%m-%d')
                elif hasattr(value, '__str__'):
                    data[field_name] = str(value)
            return JsonResponse({'status': 'success', 'data': data})
            
        return super().get(request, *args, **kwargs)

# Представления для групп оборудования
class BaseGroupMixin:
    model = EquipmentGroup
    success_url = reverse_lazy('equipment:group-list')

class EquipmentGroupListView(LoginRequiredMixin, ListView):
    model = EquipmentGroup
    template_name = 'equipment/groups/equipment_group_list.html'
    context_object_name = 'groups'
    paginate_by = 30
    ordering = ['-id']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['available_columns'] = {'name': 'Название','description': 'Описание','equipment_count': 'Количество оборудования'}
        return context
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Поиск по всем полям
        search_query = self.request.GET.get('search')
        if search_query:
            q_objects = Q()
            fields = [field.name for field in EquipmentGroup._meta.fields]
            for field in fields:
                try:
                    q_objects |= Q(**{f"{field}__icontains": search_query})
                except Exception as e:
                    print(f"Ошибка при поиске по полю {field}: {str(e)}")
                    continue
            queryset = queryset.filter(q_objects)
        
        return queryset.distinct()   

class EquipmentGroupCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = EquipmentGroup
    form_class = EquipmentGroupForm
    template_name = 'equipment/groups/equipment_group_form.html'
    success_url = reverse_lazy('equipment:group-list')
    success_message = "Группа оборудования успешно создана"

class EquipmentGroupUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = EquipmentGroup
    form_class = EquipmentGroupForm
    template_name = 'equipment/groups/equipment_group_form.html'
    success_url = reverse_lazy('equipment:group-list')
    success_message = "Группа оборудования успешно обновлена"

class EquipmentGroupDeleteView(LoginRequiredMixin, DeleteView):
    model = EquipmentGroup
    success_url = reverse_lazy('equipment:group-list')
    
    def delete(self, request, *args, **kwargs):
        try:
            self.object = self.get_object()
            self.object.delete()
            messages.success(request, 'Группа оборудования успешно удалена')
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'status': 'success'})
            return super().delete(request, *args, **kwargs)
        except Exception as e:
            messages.error(request, f'Ошибка при удалении: {str(e)}')
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
            return self.render_to_response(self.get_context_data())

    def get(self, request, *args, **kwargs):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success'})
        return self.delete(request, *args, **kwargs)

class EquipmentGroupDetailView(LoginRequiredMixin, DetailView):
    model = EquipmentGroup
    template_name = 'equipment/groups/equipment_group_detail.html'
    context_object_name = 'group'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        group = self.get_object()
        
        # Фильтруем оборудование по типам
        context.update({
            'si_equipment': group.equipment.filter(equipment_type='СИ'),
            'io_equipment': group.equipment.filter(equipment_type='ИО'),
            'vo_equipment': group.equipment.filter(equipment_type='ВО'),
        })
        
        return context

# Функции для работы с оборудованием
@login_required
def save_columns(request):
    if request.method == 'POST':
        selected_columns = request.POST.getlist('columns')
        referer = request.META.get('HTTP_REFERER', '')

        # Определяем, с какой страницы пришел запрос
        if 'equipment' in referer:
            request.session['equipment_columns'] = selected_columns
        elif 'equipment' in referer:
            request.session['equipment_columns'] = selected_columns

        messages.success(request, 'Настройки отображения сохранены')
        return redirect(referer)

@login_required
def duplicate_equipment(request, pk, equipment_type):
    source_equipment = get_object_or_404(Equipment, pk=pk)
    source_equipment.pk = None
    source_equipment.name = f"{source_equipment.name} (копия)"
    source_equipment.save()

    messages.success(request, "Элемент успешно скопирован")
    return redirect('equipment:list')


@login_required
def delete_all_equipment(request):
    if request.method == 'POST':
        deleted_count = Equipment.objects.all().delete()[0]
        messages.success(request, f'Удалено записей: {deleted_count}')
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error'}, status=400)

@login_required
def import_equipment(request):
   if request.method == 'POST':
       form = CSVImportForm(request.POST, request.FILES)
       if form.is_valid():
           try:
               file = request.FILES['csv_file']
               file_extension = file.name.split('.')[-1].lower()
               stats = {'updated': 0, 'created': 0, 'skipped': 0}
               fields = [f.name for f in Equipment._meta.fields]

               # Обработка Excel файла
               if file_extension in ['xlsx', 'xls']:
                   wb = openpyxl.load_workbook(file)
                   ws = wb.active
                   headers = {cell.value: idx for idx, cell in enumerate(ws[1])}
                   
                   for row in ws.iter_rows(min_row=2):
                       try:
                           process_row(row, headers, fields, stats)
                       except Exception as e:
                           print(f"Row processing error: {e}")
                           stats['skipped'] += 1

               # Обработка CSV файла            
               elif file_extension == 'csv':
                   import csv
                   from io import TextIOWrapper
                   
                   csv_file = TextIOWrapper(file.file, encoding='utf-8-sig')
                   reader = csv.reader(csv_file)
                   headers = {col: idx for idx, col in enumerate(next(reader))}
                   
                   for row in reader:
                       try:
                           row_data = [cell for cell in row]
                           process_row(row_data, headers, fields, stats, is_csv=True)
                       except Exception as e:
                           print(f"Row processing error: {e}")
                           stats['skipped'] += 1
               else:
                   messages.error(request, 'Неподдерживаемый формат файла. Загрузите файл .csv, .xlsx или .xls')
                   return redirect('equipment:equipment-list')

               messages.success(
                   request, 
                   f'Импорт завершен: обновлено {stats["updated"]}, '
                   f'создано {stats["created"]}, пропущено {stats["skipped"]} записей'
               )

           except Exception as e:
               messages.error(request, f'Ошибка при импорте: {str(e)}')
           
           return redirect('equipment:list')
   else:
       form = CSVImportForm()

   example_header = ','.join([f.name for f in Equipment._meta.fields])
   return render(request, 'equipment/import_csv.html', {
       'form': form,
       'example_header': example_header
   })

def process_row(row, headers, fields, stats, is_csv=False):
   """Обработка одной строки данных"""
   cleaned_data = {}
   
   # Получение ID зависит от типа файла
   if is_csv:
       item_id = row[headers.get('id', 0)] if 'id' in headers else None
   else:
       item_id = row[headers.get('id', 0)].value if 'id' in headers else None
   
   # Обработка полей
   for field in fields:
       if field in headers and field != 'id':
           if is_csv:
               value = row[headers[field]]
           else:
               value = row[headers[field]].value
               
           if value:
               if isinstance(value, str):
                   value = value.strip()
               if field in ['data_poverk', 'srok_poverk']:
                   if isinstance(value, datetime):
                       value = value.date()
                   elif isinstance(value, str):
                       try:
                           value = datetime.strptime(value, '%Y-%m-%d').date()
                       except ValueError:
                           try:
                               value = datetime.strptime(value, '%d.%m.%Y').date()
                           except ValueError:
                               value = None
           cleaned_data[field] = value

   # Создание или обновление записи
   if item_id and Equipment.objects.filter(id=item_id).exists():
       equipment = Equipment.objects.get(id=item_id)
       for field, value in cleaned_data.items():
           setattr(equipment, field, value)
       equipment.save()
       stats['updated'] += 1
   else:
       Equipment.objects.create(**cleaned_data)
       stats['created'] += 1

@login_required 
def export_equipment(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    
    fields = [field.name for field in Equipment._meta.fields]
    ws.append(fields)

    items = Equipment.objects.values_list(*fields)
    for item in items:
        ws.append(['' if value is None else value for value in item])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="equipment.xlsx"'
    wb.save(response)
    
    return response

@login_required
def bulk_delete_equipment(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        ids = data.get('ids', [])
        deleted_count = Equipment.objects.filter(id__in=ids).delete()[0]
        messages.success(request, f'Удалено элементов: {deleted_count}')
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error'}, status=400)

@login_required
def bulk_duplicate_equipment(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        ids = data.get('ids', [])
        
        duplicated_count = 0
        for equipment_id in ids:
            try:
                equipment = Equipment.objects.get(id=equipment_id)
                equipment.pk = None
                equipment.name = f"{equipment.name} (копия)"
                equipment.save()
                duplicated_count += 1
            except Equipment.DoesNotExist:
                continue
        
        messages.success(request, f'Скопировано элементов: {duplicated_count}')
        return JsonResponse({'status': 'success'})
    
    return JsonResponse({'status': 'error'}, status=400)

class DuplicateEquipmentView(LoginRequiredMixin, View):
    def post(self, request, pk, *args, **kwargs):
        return self.duplicate(request, pk)

    def get(self, request, pk, *args, **kwargs):
        return self.duplicate(request, pk)

    def duplicate(self, request, pk):
        source_equipment = get_object_or_404(Equipment, pk=pk)
        source_equipment.pk = None
        source_equipment.name = f"{source_equipment.name} (копия)"
        source_equipment.save()

        messages.success(request, "Элемент успешно скопирован")
        return redirect('equipment:list')
    
class DuplicateGroupEquipmentView(LoginRequiredMixin, View):
    def post(self, request, pk, *args, **kwargs):
        return self.duplicate(request, pk)

    def get(self, request, pk, *args, **kwargs):
        return self.duplicate(request, pk)

    def duplicate(self, request, pk):
        original_group = get_object_or_404(EquipmentGroup, pk=pk)
        new_group = EquipmentGroup.objects.create(
            name=f"Копия - {original_group.name}",
            conditions=original_group.conditions
        )
        new_group.equipment.set(original_group.equipment.all())
        messages.success(request, "Группа оборудования успешно скопирована")
        return redirect('equipment:group-list')